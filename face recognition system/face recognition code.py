import cv2
import numpy as np
import os
from datetime import date
import openpyxl

# Get current folder path
CurrentFolder = os.getcwd()

# Define image paths (we'll skip face recognition processing)
image_paths = {
    "Saniya": os.path.join(CurrentFolder, 'saniya.jpeg'),
    "Akshata": os.path.join(CurrentFolder, 'akshata.jpg'),
    "Mahuri": os.path.join(CurrentFolder, 'madhuri.jpg'),
    "Ishita": os.path.join(CurrentFolder, 'ishita.jpeg')
}

# Get a reference to the webcam
video_capture = cv2.VideoCapture(0)

if not video_capture.isOpened():
    print("Error: Could not open webcam.")
    exit()

# Load the Haar Cascade for face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Create or load the workbook
attendance_file = 'attendance.xlsx'
if os.path.exists(attendance_file):
    wb = openpyxl.load_workbook(attendance_file)
    sheet1 = wb.active
else:
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.append(['Name', 'Date'])

row = sheet1.max_row
already_attendance_taken = ""

def save_workbook():
    try:
        wb.save(attendance_file)
        print("Workbook saved successfully.")
    except Exception as e:
        print(f"Error saving workbook: {e}")

while True:
    # Grab a single frame of video
    ret, frame = video_capture.read()
    if not ret:
        print("Failed to capture image")
        break

    # Convert the image to grayscale for face detection
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Detect faces in the grayscale image
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    # Draw rectangles around the detected faces
    for (x, y, w, h) in faces:
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
        # For simplicity, mark attendance for every detected face as "Unknown"
        name = "Unknown"
        if already_attendance_taken != name:
            sheet1.append([name, "Present"])
            row += 1
            print("Attendance taken for", name)
            already_attendance_taken = name

    # Display the resulting image
    cv2.imshow('Video', frame)

    # Check if the user pressed 'q' to quit
    if cv2.waitKey(1) & 0xFF == ord('q'):
        print("Data saving...")
        save_workbook()
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
